"""
Outlook email automation for completed Lexis Cite extractions.
"""

from __future__ import annotations

from collections import Counter
from copy import copy
from dataclasses import dataclass
import os
from pathlib import Path
import re
import tempfile
from typing import Any
import sys

import openpyxl


EMAIL_SUBJECT = "RE: PLR000/CCA001 for Conversion"
WD_AUTO_FIT_CONTENT = 1
WD_AUTO_FIT_WINDOW = 2
EMAIL_ADDRESS_PATTERN = re.compile(
    r"[A-Z0-9._%+\-]+@[A-Z0-9.\-]+\.[A-Z]{2,}",
    re.IGNORECASE,
)

class NoSuccessfulExtractionsError(RuntimeError):
    """Raised when there are no successful Lexis Cite extractions to email."""


@dataclass(frozen=True)
class ExtractionResultRow:
    """Represents a single row from the extraction workbook."""

    excel_row: int
    lni: str
    filename: str
    court: str
    decided_date: Any
    lexis_cite: str
    document_type: str


@dataclass(frozen=True)
class ExtractionEmailSummary:
    """Summary used to compose the Outlook email."""

    intro_text: str
    success_rows: list[ExtractionResultRow]
    unavailable_counts: dict[str, int]
    max_excel_row: int

    @property
    def success_count(self) -> int:
        """Return the number of rows with a successfully extracted Lexis Cite."""
        return len(self.success_rows)

    @property
    def unavailable_count(self) -> int:
        """Return the number of rows still marked as unavailable."""
        return sum(self.unavailable_counts.values())


def _log(logger: Any | None, message: str) -> None:
    """Write a log entry when a logger is available."""
    if logger:
        logger.log(message)


def _clean_text(value: Any) -> str:
    """Normalize workbook text values for comparisons and display."""
    if value is None:
        return ""
    return str(value).replace("\xa0", " ").strip()


def _document_type_from_court(court: str) -> str:
    """Map the COURT column to the document family used in the email text."""
    normalized = _clean_text(court).upper()

    if "CCA001" in normalized:
        return "CCA001"
    if "PLR000" in normalized:
        return "PLR000"

    return _clean_text(court) or "PLR000/CCA001"


def _ordered_document_types(document_types: set[str]) -> list[str]:
    """Return document types in a stable, human-friendly order."""
    preferred_order = ["PLR000", "CCA001"]
    ordered = [doc_type for doc_type in preferred_order if doc_type in document_types]
    ordered.extend(sorted(document_types - set(preferred_order)))
    return ordered


def _document_label(document_types: set[str]) -> str:
    """Build the PLR000/CCA001 label portion of the intro sentence."""
    if not document_types:
        return "PLR000/CCA001"

    ordered = _ordered_document_types(document_types)

    if len(ordered) == 1:
        return ordered[0]

    if set(ordered) == {"PLR000", "CCA001"}:
        return "PLR000/CCA001"

    return "/".join(ordered)


def _format_count_groups(counts: dict[str, int]) -> str:
    """Render grouped counts like '2 PLR000 documents and 1 CCA001 document'."""
    groups: list[str] = []

    for document_type in _ordered_document_types(set(counts)):
        count = counts[document_type]
        noun = "document" if count == 1 else "documents"
        groups.append(f"{count} {document_type} {noun}")

    if not groups:
        return ""

    if len(groups) == 1:
        return groups[0]

    return ", ".join(groups[:-1]) + f" and {groups[-1]}"


def prepare_extraction_email(excel_path: Path) -> ExtractionEmailSummary:
    """
    Read the saved workbook and build the dynamic email summary.

    Args:
        excel_path: Path to the saved extraction workbook.

    Returns:
        ExtractionEmailSummary with counts and table metadata.

    Raises:
        NoSuccessfulExtractionsError: If no rows were successfully extracted.
    """
    workbook = openpyxl.load_workbook(excel_path, data_only=True)

    try:
        worksheet = workbook.active
        success_rows: list[ExtractionResultRow] = []
        unavailable_counts: Counter[str] = Counter()
        max_excel_row = 1

        for row_index in range(2, worksheet.max_row + 1):
            row_values = [
                worksheet[f"{column}{row_index}"].value
                for column in ("A", "B", "C", "D", "E")
            ]

            if not any(_clean_text(value) for value in row_values):
                continue

            max_excel_row = row_index

            lni, filename, court, decided_date, lexis_cite = row_values
            cleaned_court = _clean_text(court)
            cleaned_cite = _clean_text(lexis_cite)

            result_row = ExtractionResultRow(
                excel_row=row_index,
                lni=_clean_text(lni),
                filename=_clean_text(filename),
                court=cleaned_court,
                decided_date=decided_date,
                lexis_cite=cleaned_cite,
                document_type=_document_type_from_court(cleaned_court),
            )

            if cleaned_cite and cleaned_cite.lower() != "not available":
                success_rows.append(result_row)
            else:
                unavailable_counts[result_row.document_type] += 1

        if not success_rows:
            raise NoSuccessfulExtractionsError(
                "Extraction completed, but no Lexis Cites were successfully extracted, "
                "so no Outlook email was sent."
            )

        document_types = {row.document_type for row in success_rows}
        success_count = len(success_rows)
        availability_verb = "is" if success_count == 1 else "are"
        intro_text = (
            f"The {success_count} auto-collected {_document_label(document_types)} "
            f"{'document' if success_count == 1 else 'documents'} below {availability_verb} already "
            "available online"
        )

        if unavailable_counts:
            total_unavailable = sum(unavailable_counts.values())
            verb = "is" if total_unavailable == 1 else "are"
            intro_text += f", while {_format_count_groups(dict(unavailable_counts))} {verb} stuck in keying."
        else:
            intro_text += "."

        return ExtractionEmailSummary(
            intro_text=intro_text,
            success_rows=success_rows,
            unavailable_counts=dict(unavailable_counts),
            max_excel_row=max_excel_row,
        )
    finally:
        workbook.close()


def _build_email_table_workbook(excel_path: Path, summary: ExtractionEmailSummary) -> Path:
    """
    Create a temporary workbook copy for Outlook pasting.

    The copy keeps all extracted rows in the table and rewrites unavailable
    entries in Column E as italicized "Stuck in Keying" so the receiving team
    can follow up.
    """
    file_descriptor, temp_path_str = tempfile.mkstemp(
        prefix="plr000_cca001_email_",
        suffix=".xlsx",
    )
    os.close(file_descriptor)
    temp_path = Path(temp_path_str)

    workbook = openpyxl.load_workbook(excel_path)

    try:
        worksheet = workbook.active

        for row_index in range(2, summary.max_excel_row + 1):
            cell = worksheet[f"E{row_index}"]
            if _clean_text(cell.value).lower() != "not available":
                continue

            cell.value = "Stuck in Keying"
            italic_font = copy(cell.font)
            italic_font.italic = True
            cell.font = italic_font

        workbook.save(temp_path)
        return temp_path
    finally:
        workbook.close()


def _create_outlook_reply(
    outlook_app: Any,
    source_message_entry_id: str | None,
    source_message_store_id: str | None,
    fallback_to: str | None,
    fallback_cc: str | None,
    override_to: str | None,
    override_cc: str | None,
    logger: Any | None = None,
) -> Any:
    """
    Create the outbound Outlook item using the live source thread when possible.

    Falling back to the source message's current To/CC values keeps recipients
    dynamic even if ReplyAll cannot be resolved for some reason.
    """
    if override_to:
        mail_item = outlook_app.CreateItem(0)
        _apply_override_recipients(
            mail_item=mail_item,
            override_to=override_to,
            override_cc=override_cc,
            logger=logger,
        )
        mail_item.Subject = EMAIL_SUBJECT
        _log(logger, "Created Outlook email with manually overridden recipients")
        return mail_item

    if source_message_entry_id:
        try:
            namespace = outlook_app.GetNamespace("MAPI")
            source_item = namespace.GetItemFromID(
                source_message_entry_id,
                source_message_store_id or None,
            )
            reply_item = source_item.ReplyAll()
            _log(logger, "Created Outlook reply from the source email thread")
            return reply_item
        except Exception as e:
            _log(logger, f"ReplyAll fallback to dynamic To/CC will be used: {e}")

    mail_item = outlook_app.CreateItem(0)
    mail_item.To = fallback_to or ""
    mail_item.CC = fallback_cc or ""
    mail_item.Subject = EMAIL_SUBJECT
    _log(logger, "Created new Outlook email using dynamic recipients from the source email")
    return mail_item


def _move_cursor_to_top(selection: Any) -> None:
    """Move the Word editor selection to the start of the message body."""
    try:
        selection.HomeKey(Unit=6)
    except Exception:
        pass


def _split_override_recipients(raw_recipients: str | None) -> list[str]:
    """Extract clean SMTP-style addresses from Outlook-copied recipient text."""
    if not raw_recipients:
        return []

    normalized = str(raw_recipients).replace("\r", "\n")
    raw_segments = [
        segment.strip()
        for segment in re.split(r"[;\n]+", normalized)
        if segment.strip()
    ]

    if not raw_segments:
        raw_segments = [normalized.strip()]

    cleaned_addresses: list[str] = []
    seen_addresses: set[str] = set()

    for segment in raw_segments:
        candidate_addresses: list[str] = []
        bracket_matches = re.findall(r"<([^<>]+)>", segment)

        if bracket_matches:
            for bracket_value in bracket_matches:
                candidate_addresses.extend(EMAIL_ADDRESS_PATTERN.findall(bracket_value))
        else:
            candidate_addresses.extend(EMAIL_ADDRESS_PATTERN.findall(segment))

        for candidate in candidate_addresses:
            normalized_candidate = _clean_text(candidate).strip("\"'<>")
            if not normalized_candidate:
                continue

            dedupe_key = normalized_candidate.lower()
            if dedupe_key in seen_addresses:
                continue

            seen_addresses.add(dedupe_key)
            cleaned_addresses.append(normalized_candidate)

    return cleaned_addresses


def _apply_override_recipients(
    mail_item: Any,
    override_to: str | None,
    override_cc: str | None,
    logger: Any | None = None,
) -> None:
    """Normalize manual override recipients and assign them to Outlook."""
    to_entries = _split_override_recipients(override_to)
    cc_entries = _split_override_recipients(override_cc)

    if not to_entries:
        raise RuntimeError("Manual recipient override requires at least one valid To recipient.")

    mail_item.To = "; ".join(to_entries)
    mail_item.CC = "; ".join(cc_entries)

    _log(
        logger,
        "Normalized manual override recipients to "
        f"To='{mail_item.To}' CC='{mail_item.CC}'",
    )


def _resolve_mail_recipients(
    mail_item: Any,
    override_to: str | None,
    override_cc: str | None,
    fallback_to: str | None,
    fallback_cc: str | None,
) -> tuple[str, str]:
    """Capture recipient strings before Outlook moves the item on send."""
    resolved_to = override_to or fallback_to or ""
    resolved_cc = override_cc or fallback_cc or ""

    try:
        resolved_to = getattr(mail_item, "To", resolved_to) or resolved_to
    except Exception:
        pass

    try:
        resolved_cc = getattr(mail_item, "CC", resolved_cc) or resolved_cc
    except Exception:
        pass

    return resolved_to, resolved_cc


def _format_pasted_outlook_table(word_editor: Any, logger: Any | None = None) -> None:
    """Auto-fit the pasted Outlook table so long values wrap cleanly."""
    try:
        table = word_editor.Tables(1)
    except Exception as exc:
        _log(logger, f"Could not find the pasted Outlook table to format: {exc}")
        return

    try:
        table.AllowAutoFit = True
    except Exception:
        pass

    try:
        # First fit to content, then constrain back to the email window width.
        table.AutoFitBehavior(WD_AUTO_FIT_CONTENT)
        table.AutoFitBehavior(WD_AUTO_FIT_WINDOW)
        _log(logger, "Applied Outlook table auto-fit formatting")
    except Exception as exc:
        _log(logger, f"Could not auto-fit the Outlook table after paste: {exc}")


def send_extraction_email(
    excel_path: Path,
    logger: Any | None = None,
    source_message_entry_id: str | None = None,
    source_message_store_id: str | None = None,
    fallback_to: str | None = None,
    fallback_cc: str | None = None,
    override_to: str | None = None,
    override_cc: str | None = None,
) -> ExtractionEmailSummary:
    """
    Create and send the Outlook email with the successful extraction table.

    Args:
        excel_path: Path to the saved extraction workbook.
        logger: Optional logger for status updates.
        source_message_entry_id: Outlook EntryID for the source message/thread.
        source_message_store_id: Outlook StoreID for the source message/thread.
        fallback_to: Dynamic To recipients captured from the source email.
        fallback_cc: Dynamic CC recipients captured from the source email.
        override_to: Override To recipients for safe test sends.
        override_cc: Override CC recipients for safe test sends.

    Returns:
        The prepared email summary that was sent.
    """
    if sys.platform != "win32":
        raise RuntimeError("Outlook automation is only supported on Windows.")

    summary = prepare_extraction_email(excel_path)

    try:
        import pythoncom
        import win32com.client
    except ImportError as exc:
        raise RuntimeError(
            "pywin32 is required for Outlook automation but is not installed."
        ) from exc

    if not override_to and not any((source_message_entry_id, fallback_to, fallback_cc)):
        raise RuntimeError(
            "No Outlook reply context is available for the automatic email send. "
            "Use Outlook Email source, or provide override recipients in Recipients or Developer Mode."
        )

    excel_app = None
    workbook = None
    email_table_workbook_path = None

    pythoncom.CoInitialize()

    try:
        _log(logger, "Preparing Outlook email with successful extraction rows")
        email_table_workbook_path = _build_email_table_workbook(excel_path, summary)

        excel_app = win32com.client.DispatchEx("Excel.Application")
        excel_app.Visible = False
        excel_app.DisplayAlerts = False

        workbook = excel_app.Workbooks.Open(str(email_table_workbook_path), ReadOnly=True)
        worksheet = workbook.Worksheets(1)

        if worksheet.AutoFilterMode:
            worksheet.AutoFilterMode = False

        table_range = worksheet.Range(f"A1:E{summary.max_excel_row}")
        table_range.Copy()

        outlook_app = win32com.client.Dispatch("Outlook.Application")
        mail_item = _create_outlook_reply(
            outlook_app=outlook_app,
            source_message_entry_id=source_message_entry_id,
            source_message_store_id=source_message_store_id,
            fallback_to=fallback_to,
            fallback_cc=fallback_cc,
            override_to=override_to,
            override_cc=override_cc,
            logger=logger,
        )
        mail_item.Display()

        word_editor = mail_item.GetInspector.WordEditor
        selection = word_editor.Application.Selection
        _move_cursor_to_top(selection)
        selection.TypeText("Hi All,")
        selection.TypeParagraph()
        selection.TypeParagraph()
        selection.TypeText(summary.intro_text)
        selection.TypeParagraph()
        selection.TypeParagraph()
        selection.Range.PasteExcelTable(False, False, False)
        _format_pasted_outlook_table(word_editor, logger)

        resolved_to, resolved_cc = _resolve_mail_recipients(
            mail_item=mail_item,
            override_to=override_to,
            override_cc=override_cc,
            fallback_to=fallback_to,
            fallback_cc=fallback_cc,
        )

        try:
            mail_item.Save()
        except Exception as exc:
            _log(logger, f"Could not save Outlook draft before send: {exc}")

        mail_item.Send()
        _log(
            logger,
            "Outlook email sent successfully using "
            f"{'manual override recipients' if override_to else 'dynamic recipients from the source email'}, "
            f"To='{resolved_to}', CC='{resolved_cc}', "
            "with the full extraction table, "
            f"including {summary.unavailable_count} follow-up "
            f"{'item' if summary.unavailable_count == 1 else 'items'} marked as Stuck in Keying",
        )
        return summary

    finally:
        if workbook is not None:
            try:
                if workbook.Worksheets(1).AutoFilterMode:
                    workbook.Worksheets(1).AutoFilterMode = False
            except Exception:
                pass

            try:
                workbook.Close(False)
            except Exception:
                pass

        if excel_app is not None:
            try:
                excel_app.CutCopyMode = False
            except Exception:
                pass

            try:
                excel_app.Quit()
            except Exception:
                pass

        if email_table_workbook_path is not None:
            try:
                Path(email_table_workbook_path).unlink(missing_ok=True)
            except Exception:
                pass

        pythoncom.CoUninitialize()
