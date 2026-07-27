"""
Excel file handling module for reading and writing Excel files.
"""
import openpyxl
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import subprocess
import sys
import os


class ExcelHandler:
    """Handles Excel file operations."""

    COLUMN_HEADERS = {
        'A': 'LNI',
        'B': 'FILENAME',
        'C': 'COURT',
        'D': 'DECIDED DATE',
        'E': 'LEXIS CITE'
    }
    DEFAULT_HEADER_FILL_COLOR = "A9D08E"
    HEADER_COLOR_PRESETS = (
        ("Sage Green", "A9D08E"),
        ("Moss Green", "70AD47"),
        ("Emerald", "00B050"),
        ("Seafoam", "C6E0B4"),
        ("Teal", "4BACC6"),
        ("Aqua", "9FD9D5"),
        ("Sky Blue", "5B9BD5"),
        ("Navy", "4472C4"),
        ("Midnight", "1F4E78"),
        ("Lavender", "B4A7D6"),
        ("Plum", "8064A2"),
        ("Rose", "E6B8B7"),
        ("Crimson", "C0504D"),
        ("Coral", "F4B183"),
        ("Orange", "ED7D31"),
        ("Amber", "FFC000"),
        ("Gold", "FFD966"),
        ("Stone", "D9D9D9"),
        ("Slate", "A5A5A5"),
        ("Charcoal", "595959"),
    )

    def __init__(self, file_path: Path = None, header_fill_color: str | None = None):
        """
        Initialize the Excel handler.
        
        Args:
            file_path: Path to Excel file (optional)
        """
        self.file_path = file_path
        self.workbook = None
        self.worksheet = None
        self.header_fill_color = self.normalize_header_fill_color(header_fill_color)

    @classmethod
    def normalize_header_fill_color(cls, color_value: str | None) -> str:
        """Normalize a color string into a six-character Excel-friendly hex value."""
        if not color_value:
            return cls.DEFAULT_HEADER_FILL_COLOR

        normalized = str(color_value).strip().lstrip("#").upper()
        if len(normalized) == 8:
            normalized = normalized[-6:]

        valid_hex = "0123456789ABCDEF"
        if len(normalized) != 6 or any(character not in valid_hex for character in normalized):
            return cls.DEFAULT_HEADER_FILL_COLOR

        return normalized

    @staticmethod
    def _header_font_color(fill_color: str) -> str:
        """Choose a readable header font color based on the fill brightness."""
        red = int(fill_color[0:2], 16)
        green = int(fill_color[2:4], 16)
        blue = int(fill_color[4:6], 16)
        brightness = (red * 299) + (green * 587) + (blue * 114)
        return "000000" if brightness >= 140000 else "FFFFFF"

    @staticmethod
    def _normalize_row_text(value) -> str:
        """Normalize worksheet text for header-row detection."""
        return " ".join(str(value or "").strip().lower().split())

    def _is_header_like_source_row(self, row: int) -> bool:
        """Detect imported table headers that accidentally landed in the data area."""
        values = [
            self._normalize_row_text(self.worksheet[f"{column}{row}"].value)
            for column in ("A", "B", "C", "D")
        ]
        return values in (
            ["lni", "filename", "court", "decided date"],
            ["lni", "file name", "court code", "decided date"],
        )

    def _autofit_columns(self, worksheet):
        """
        Auto-fit column widths based on content.
        
        Args:
            worksheet: The worksheet to auto-fit
        """
        max_row = worksheet.max_row
        
        for col_letter in self.COLUMN_HEADERS.keys():
            max_length = 0
            
            # Check header
            header_cell = worksheet[f"{col_letter}1"]
            if header_cell.value:
                max_length = max(max_length, len(str(header_cell.value)))
            
            # Check all data cells in the column
            for row in range(2, max_row + 1):
                cell = worksheet[f"{col_letter}{row}"]
                if cell.value:
                    # Get the length of the cell value
                    cell_length = len(str(cell.value))
                    max_length = max(max_length, cell_length)
            
            # Set column width (add some padding: +2 for better readability)
            # Excel column width is in character units, with a max of ~255
            column_width = min(max_length + 2, 255)
            # Set minimum width to ensure headers are visible
            column_width = max(column_width, len(self.COLUMN_HEADERS[col_letter]) + 2)
            worksheet.column_dimensions[col_letter].width = column_width
    
    def _apply_formatting(self, worksheet):
        """
        Apply formatting to the worksheet (headers and data).
        
        Args:
            worksheet: The worksheet to format
        """
        # Auto-fit columns first
        self._autofit_columns(worksheet)
        
        # Define border style
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        header_fill_color = self.normalize_header_fill_color(self.header_fill_color)
        header_fill = PatternFill(
            start_color=header_fill_color,
            end_color=header_fill_color,
            fill_type="solid",
        )
        header_font = Font(
            bold=True,
            color=self._header_font_color(header_fill_color),
        )
        
        # Define center alignment
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        # Get the data range
        max_row = worksheet.max_row
        max_col = len(self.COLUMN_HEADERS)
        
        # Format headers (row 1)
        for col_idx, col_letter in enumerate(self.COLUMN_HEADERS.keys(), 1):
            cell = worksheet[f"{col_letter}1"]
            cell.font = header_font
            cell.alignment = center_alignment
            cell.fill = header_fill
            cell.border = thin_border
        
        # Format data rows (row 2 onwards)
        for row in range(2, max_row + 1):
            for col_letter in self.COLUMN_HEADERS.keys():
                cell = worksheet[f"{col_letter}{row}"]
                cell.alignment = center_alignment
                cell.border = thin_border
    
    def create_template(self, run_folder: Path = None) -> Path:
        """
        Create a new formatted Excel template file in run subfolder.
        
        Args:
            run_folder: Run folder path where to save the template (optional)
            
        Returns:
            Path to the created Excel file
        """
        from utils.file_manager import FileManager
        file_manager = FileManager()
        
        if run_folder is None:
            run_folder = file_manager.create_run_folder()
        
        # Generate output filename using same convention
        output_path = file_manager.get_output_excel_path(run_folder)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        
        # Set headers
        for col, header in self.COLUMN_HEADERS.items():
            cell = ws[f"{col}1"]
            cell.value = header
        
        # Apply formatting (includes auto-fit columns)
        self._apply_formatting(ws)
        
        wb.save(output_path)
        self.file_path = output_path
        return output_path
    
    def open_excel_file(self, file_path: Path):
        """
        Open an existing Excel file.
        
        Args:
            file_path: Path to the Excel file
        """
        self.file_path = file_path
        self.workbook = openpyxl.load_workbook(file_path)
        self.worksheet = self.workbook.active
    
    def read_lni_data(self) -> list:
        """
        Read LNI data from Column A.
        
        Returns:
            List of tuples: (row_number, lni_value)
        """
        if not self.worksheet:
            raise ValueError("No Excel file loaded")
        
        lni_data = []
        for row in range(2, self.worksheet.max_row + 1):
            if self._is_header_like_source_row(row):
                continue
            cell_value = self.worksheet[f'A{row}'].value
            if cell_value and str(cell_value).strip():
                lni_data.append((row, str(cell_value).strip()))
        
        return lni_data

    def clear_data_rows(self):
        """
        Remove all existing data rows while preserving the header row.
        """
        if not self.worksheet:
            raise ValueError("No Excel file loaded")

        if self.worksheet.max_row > 1:
            self.worksheet.delete_rows(2, self.worksheet.max_row - 1)

    def populate_source_rows(self, rows: list[dict[str, str]]):
        """
        Populate the worksheet with imported source rows.

        Args:
            rows: List of dictionaries keyed by the Excel header names.
        """
        if not self.worksheet:
            raise ValueError("No Excel file loaded")

        self.clear_data_rows()

        target_row = 2
        for row_data in rows:
            normalized_values = {
                header: str(row_data.get(header, "") or "").strip()
                for header in self.COLUMN_HEADERS.values()
            }

            if not any(normalized_values.values()):
                continue
            if [
                self._normalize_row_text(normalized_values["LNI"]),
                self._normalize_row_text(normalized_values["FILENAME"]),
                self._normalize_row_text(normalized_values["COURT"]),
                self._normalize_row_text(normalized_values["DECIDED DATE"]),
            ] in (
                ["lni", "filename", "court", "decided date"],
                ["lni", "file name", "court code", "decided date"],
            ):
                continue

            for column_letter, header in self.COLUMN_HEADERS.items():
                self.worksheet[f"{column_letter}{target_row}"].value = normalized_values[header]

            target_row += 1

        if target_row == 2:
            raise ValueError("No non-empty source rows were available to populate the workbook")
    
    def write_lexis_cite(self, row: int, lexis_cite: str):
        """
        Write Lexis Cite to Column E.
        
        Args:
            row: Row number (1-indexed)
            lexis_cite: Lexis Cite value to write
        """
        if not self.worksheet:
            raise ValueError("No Excel file loaded")
        
        self.worksheet[f'E{row}'].value = lexis_cite
    
    def save(self, output_path: Path = None):
        """
        Save the Excel file with formatting applied.
        
        Args:
            output_path: Path where to save (optional, uses original path if not provided)
            
        Raises:
            ValueError: If workbook or path is not set
            PermissionError: If file is locked or cannot be written
            OSError: If path is invalid or file system error occurs
        """
        if not self.workbook:
            raise ValueError("No workbook to save")
        
        save_path = output_path or self.file_path
        if not save_path:
            raise ValueError("No file path specified")
        
        # Convert to string and validate path
        save_path_str = str(save_path)
        
        # Check for invalid characters in filename (Windows)
        invalid_chars = ['<', '>', ':', '"', '/', '\\', '|', '?', '*']
        filename = save_path.name
        if any(char in filename for char in invalid_chars):
            raise ValueError(f"Filename contains invalid characters: {filename}")
        
        # Apply formatting before saving
        if self.worksheet:
            self._apply_formatting(self.worksheet)
        
        try:
            self.workbook.save(save_path_str)
        except PermissionError as e:
            raise PermissionError(f"Cannot save file - it may be open in another application: {save_path_str}") from e
        except OSError as e:
            raise OSError(f"File system error saving to {save_path_str}: {str(e)}") from e
    
    def open_file(self, file_path: Path):
        """
        Open the Excel file in the default application.
        
        Args:
            file_path: Path to the Excel file
        """
        try:
            if sys.platform == 'win32':
                os.startfile(file_path)
            elif sys.platform == 'darwin':
                subprocess.run(['open', str(file_path)])
            else:
                subprocess.run(['xdg-open', str(file_path)])
        except Exception as e:
            print(f"Error opening file: {e}")

